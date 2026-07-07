import io

from flask import Blueprint, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, ListFlowable, ListItem
)

from models import Meeting
from utils import roles_required

exports_bp = Blueprint("exports", __name__, url_prefix="/api/meetings")


def _meeting_time_label(meeting):
    start = meeting.start_time.strftime("%I:%M %p").lstrip("0") if meeting.start_time else "—"
    end = meeting.end_time.strftime("%I:%M %p").lstrip("0") if meeting.end_time else "—"
    return f"{start} – {end}"


# ---------------------------------------------------------------------------
# Feature 3: Automatic Meeting Agenda PDF
# ---------------------------------------------------------------------------

@exports_bp.route("/<int:meeting_id>/agenda-pdf", methods=["GET"])
@roles_required("admin", "super_admin")
def agenda_pdf(meeting_id):
    meeting = Meeting.query.get_or_404(meeting_id)
    sections = meeting.agenda_sections.order_by(None).all()  # already ordered via relationship

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        topMargin=1.8 * cm, bottomMargin=1.8 * cm, leftMargin=2 * cm, rightMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    college = ParagraphStyle("College", parent=styles["Title"], fontName="Times-Bold", fontSize=16, leading=19, spaceAfter=2, alignment=1)
    sub = ParagraphStyle("Sub", parent=styles["Normal"], fontName="Times-Roman", fontSize=9.5, alignment=1, spaceAfter=2)
    assoc = ParagraphStyle("Assoc", parent=styles["Normal"], fontSize=14, alignment=1, fontName="Times-Bold", spaceBefore=2, spaceAfter=14)
    doc_title = ParagraphStyle("DocTitle", parent=styles["Normal"], fontSize=13, alignment=1, fontName="Times-Bold", spaceAfter=16)
    label = ParagraphStyle("Label", parent=styles["Normal"], fontName="Times-Roman", fontSize=11, spaceAfter=4)
    body = ParagraphStyle("Body", parent=styles["Normal"], fontName="Times-Roman", fontSize=11, spaceAfter=6, leading=14)
    section_heading = ParagraphStyle("SectionHeading", parent=styles["Normal"], fontName="Times-BoldItalic", fontSize=11.5, spaceBefore=12, spaceAfter=4)
    bullet_style = ParagraphStyle("Bullet", parent=styles["Normal"], fontName="Times-Roman", fontSize=11, leading=15)

    story = []
    story.append(Paragraph("THIAGARAJAR COLLEGE OF ENGINEERING", college))
    story.append(Paragraph("(A Govt. aided Autonomous Institution Affiliated to Anna University)", sub))
    story.append(Paragraph("Civil Engineering Association", assoc))
    story.append(Paragraph("MEETING AGENDA", doc_title))

    story.append(Paragraph(f"<b><i>Date</i></b>: {meeting.meeting_date.strftime('%d %B %Y') if meeting.meeting_date else '—'}", label))
    story.append(Paragraph(f"<b><i>Time</i></b>: {_meeting_time_label(meeting)}", label))
    story.append(Paragraph(f"<b><i>Venue</i></b>: {meeting.venue or '—'}", label))
    story.append(Paragraph(f"<b><i>Agenda</i></b>: {meeting.title}", label))

    story.append(Spacer(1, 10))
    story.append(Paragraph("Purpose of Meeting", section_heading))
    purpose_text = (meeting.description or meeting.agenda or "").strip()
    if purpose_text:
        story.append(Paragraph(purpose_text, body))

    for section in sections:
        story.append(Paragraph(section.title, section_heading))
        bullets = section.bullet_points or []
        if bullets:
            story.append(ListFlowable(
                [ListItem(Paragraph(b, bullet_style)) for b in bullets],
                bulletType="bullet", start="•", leftIndent=14,
            ))

    story.append(Paragraph("Open Discussion", section_heading))
    story.append(ListFlowable(
        [ListItem(Paragraph("Suggestions and additional ideas", bullet_style))],
        bulletType="bullet", start="•", leftIndent=14,
    ))
    story.append(Spacer(1, 30))

    # Footer signatures: General Secretary / Joint Secretary on one row,
    # Staff Advisor centered on the row below (per the corrected layout —
    # NOT "Joint Secretary / Staff Advisor" paired together).
    sig_table = Table(
        [
            ["General Secretary", "Joint Secretary"],
            ["Staff Advisor", ""],
        ],
        colWidths=[8 * cm, 8 * cm],
    )
    sig_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Times-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 11),
        ("ALIGN", (0, 0), (0, 0), "LEFT"),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
        ("SPAN", (0, 1), (1, 1)),
        ("ALIGN", (0, 1), (1, 1), "CENTER"),
        ("TOPPADDING", (0, 1), (-1, 1), 24),
    ]))
    story.append(sig_table)

    doc.build(story)
    buffer.seek(0)

    filename = f"meeting-agenda-{meeting.id}.pdf"
    return send_file(buffer, mimetype="application/pdf", as_attachment=True, download_name=filename)


# ---------------------------------------------------------------------------
# Feature 6: Export options — per-meeting attendance Excel / PDF
# ---------------------------------------------------------------------------

def _attendance_rows(meeting):
    records = meeting.attendance_records.order_by(None).all()
    rows = []
    for r in records:
        rows.append([
            r.student.name if r.student else "—",
            r.student.register_number if r.student else "—",
            r.status.capitalize(),
            meeting.title,
            meeting.meeting_date.strftime("%d %b %Y") if meeting.meeting_date else "—",
        ])
    return rows


@exports_bp.route("/<int:meeting_id>/attendance-excel", methods=["GET"])
@roles_required("admin", "super_admin")
def attendance_excel(meeting_id):
    meeting = Meeting.query.get_or_404(meeting_id)
    rows = _attendance_rows(meeting)

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    headers = ["Student Name", "Register Number", "Attendance Status", "Meeting Name", "Date"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append(row)

    widths = [26, 18, 18, 28, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"attendance-{meeting.id}.xlsx"
    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=filename,
    )


@exports_bp.route("/<int:meeting_id>/attendance-pdf", methods=["GET"])
@roles_required("admin", "super_admin")
def attendance_pdf(meeting_id):
    meeting = Meeting.query.get_or_404(meeting_id)
    rows = _attendance_rows(meeting)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()

    story = [
        Paragraph(f"Attendance Report — {meeting.title}", styles["Title"]),
        Paragraph(
            f"{meeting.meeting_date.strftime('%d %B %Y') if meeting.meeting_date else '—'} · "
            f"{_meeting_time_label(meeting)} · {meeting.venue or 'Venue TBD'}",
            styles["Normal"],
        ),
        Spacer(1, 12),
    ]

    table_data = [["Student Name", "Register Number", "Status", "Meeting", "Date"]] + rows
    table = Table(table_data, repeatRows=1, colWidths=[5.5 * cm, 3.2 * cm, 2.5 * cm, 4.5 * cm, 2.5 * cm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F3A5F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5C0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F6F4")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(table)

    doc.build(story)
    buffer.seek(0)

    filename = f"attendance-{meeting.id}.pdf"
    return send_file(buffer, mimetype="application/pdf", as_attachment=True, download_name=filename)
